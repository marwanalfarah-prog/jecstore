"""Report and dashboard exports (Part I §2.2).

§2.2: *"Every report and dashboard view must be exportable (CSV, Excel, and PDF
at minimum) for offline use and accounting handoff."*

So this is one shared layer rather than per-screen export code: a screen builds a
:class:`Dataset` and hands it over, and CSV/Excel/PDF/print all follow. Adding a
new exportable report means describing its columns, nothing more.

**Arabic is the constraint that shapes every format here.**

* **CSV** is written with a UTF-8 BOM. Without it Excel on Windows opens Arabic
  as mojibake, which is exactly the accounting handoff §2.2 is for.
* **Excel** sets right-to-left sheet direction when the export is Arabic, so the
  columns read in the right order rather than merely containing Arabic text.
* **PDF** uses WeasyPrint, which Part II §7.4 recommends *because* it handles
  RTL natively. WeasyPrint needs native GTK/Pango libraries; where those are
  missing (a bare Windows dev box), :func:`pdf_available` is False and the
  screens offer the print view instead — a browser's Print → Save as PDF
  produces correct RTL output. The feature degrades; it does not disappear.
"""

from __future__ import annotations

import csv
import datetime as dt
import io
import re
from dataclasses import dataclass, field
from decimal import Decimal
from typing import Any, Literal

from fastapi import Request
from fastapi.responses import Response, StreamingResponse

from app.core.errors import AppError
from app.core.logging import get_logger

log = get_logger(__name__)

ExportFormat = Literal["csv", "xlsx", "pdf", "html"]
EXPORT_FORMATS: tuple[str, ...] = ("csv", "xlsx", "pdf")

#: Excel caps sheet names at 31 characters and forbids these.
_SHEET_FORBIDDEN = re.compile(r"[\\/*?:\[\]]")
#: Values Excel would interpret as a formula if they led a cell.
_FORMULA_LEAD = ("=", "+", "-", "@")


def _weasyprint_available() -> bool:
    """True only if WeasyPrint can actually render, not merely import.

    The package imports fine on Windows while its native libraries are missing,
    so a bare ``import weasyprint`` is not a usable check.
    """
    try:
        from weasyprint import HTML  # noqa: F401
    except Exception:  # noqa: BLE001 - any import-time failure means unusable
        return False
    try:
        from weasyprint import HTML

        HTML(string="<p>.</p>").write_pdf()
        return True
    except Exception:  # noqa: BLE001 - missing GTK/Pango surfaces here
        return False


_PDF_AVAILABLE: bool | None = None


def pdf_available() -> bool:
    """Whether PDF export can run in this environment (cached)."""
    global _PDF_AVAILABLE
    if _PDF_AVAILABLE is None:
        _PDF_AVAILABLE = _weasyprint_available()
        if not _PDF_AVAILABLE:
            log.info(
                "pdf_export_unavailable",
                extra={"reason": "WeasyPrint native libraries not present"},
            )
    return _PDF_AVAILABLE


@dataclass(slots=True)
class Column:
    """One column: where the value comes from and how it should read."""

    key: str
    label: str
    #: ``text`` | ``number`` | ``money`` | ``date`` — drives alignment and the
    #: Excel cell format, so a total sums in the spreadsheet rather than sitting
    #: there as a string.
    kind: str = "text"


@dataclass(slots=True)
class Dataset:
    """A report, ready to be written in any format."""

    title: str
    columns: list[Column]
    rows: list[dict[str, Any]]
    language: str = "en"
    #: Filters and context shown in the header — "which period is this?" is the
    #: first question anyone asks of an exported report.
    meta: dict[str, str] = field(default_factory=dict)
    #: Appended as a bold final row where present.
    totals: dict[str, Any] | None = None

    @property
    def is_rtl(self) -> bool:
        return self.language == "ar"

    def cell(self, row: dict[str, Any], column: Column) -> Any:
        return row.get(column.key)


def filename(dataset: Dataset, fmt: str) -> str:
    """The download name, which may legitimately be Arabic."""
    stem = re.sub(r"[^\w\-]+", "-", dataset.title, flags=re.UNICODE).strip("-").lower()
    stamp = dt.datetime.now(dt.timezone.utc).strftime("%Y%m%d")
    return f"{stem or 'report'}-{stamp}.{fmt}"


def content_disposition(dataset: Dataset, fmt: str) -> str:
    """Build a ``Content-Disposition`` an Arabic filename survives.

    HTTP headers are latin-1, so an Arabic name cannot go in the plain
    ``filename=`` parameter — it raises on encode. RFC 5987's ``filename*``
    carries the real UTF-8 name, with an ASCII ``filename=`` kept alongside for
    clients that ignore it.
    """
    from urllib.parse import quote

    full = filename(dataset, fmt)
    ascii_fallback = re.sub(r"[^A-Za-z0-9._-]+", "-", full).strip("-") or f"report.{fmt}"
    return (
        f'attachment; filename="{ascii_fallback}"; '
        f"filename*=UTF-8''{quote(full, safe='')}"
    )


# ---------------------------------------------------------------------------
# CSV
# ---------------------------------------------------------------------------


def _csv_safe(value: Any) -> Any:
    """Neutralise values a spreadsheet would execute as a formula.

    An exported customer note starting with ``=`` should be text, not a
    calculation — this is CSV injection, and accounting handoff is exactly where
    it would land.
    """
    if isinstance(value, str) and value.startswith(_FORMULA_LEAD):
        return "'" + value
    return value


def to_csv(dataset: Dataset) -> bytes:
    buffer = io.StringIO(newline="")
    writer = csv.writer(buffer, lineterminator="\r\n")

    writer.writerow([dataset.title])
    for key, value in dataset.meta.items():
        writer.writerow([key, value])
    if dataset.meta:
        writer.writerow([])

    writer.writerow([c.label for c in dataset.columns])
    for row in dataset.rows:
        writer.writerow([_csv_safe(_plain(dataset.cell(row, c))) for c in dataset.columns])

    if dataset.totals:
        writer.writerow([_csv_safe(_plain(dataset.totals.get(c.key, ""))) for c in dataset.columns])

    # UTF-8 BOM: without it Excel on Windows shows Arabic as mojibake.
    return b"\xef\xbb\xbf" + buffer.getvalue().encode("utf-8")


def _plain(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, Decimal):
        return str(value)
    if isinstance(value, (dt.datetime, dt.date)):
        return value.isoformat()
    if isinstance(value, bool):
        return "1" if value else "0"
    return value


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------


def to_xlsx(dataset: Dataset) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = _SHEET_FORBIDDEN.sub("-", dataset.title)[:31] or "Report"

    # Right-to-left sheet, so the columns themselves read in the right order.
    sheet.sheet_view.rightToLeft = dataset.is_rtl

    row_index = 1
    sheet.cell(row=row_index, column=1, value=dataset.title).font = Font(bold=True, size=14)
    row_index += 1

    for key, value in dataset.meta.items():
        sheet.cell(row=row_index, column=1, value=key).font = Font(italic=True)
        sheet.cell(row=row_index, column=2, value=value)
        row_index += 1
    if dataset.meta:
        row_index += 1

    header_row = row_index
    header_fill = PatternFill("solid", fgColor="2C3446")  # navy, matching §17.1
    for index, column in enumerate(dataset.columns, start=1):
        cell = sheet.cell(row=header_row, column=index, value=column.label)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="right" if dataset.is_rtl else "left")
    row_index += 1

    for row in dataset.rows:
        for index, column in enumerate(dataset.columns, start=1):
            _write_cell(sheet.cell(row=row_index, column=index), dataset.cell(row, column), column)
        row_index += 1

    if dataset.totals:
        for index, column in enumerate(dataset.columns, start=1):
            cell = sheet.cell(row=row_index, column=index)
            _write_cell(cell, dataset.totals.get(column.key), column)
            cell.font = Font(bold=True)

    # Freeze the header so a long report stays readable while scrolling.
    sheet.freeze_panes = sheet.cell(row=header_row + 1, column=1)

    for index, column in enumerate(dataset.columns, start=1):
        widest = max(
            [len(str(column.label))]
            + [len(str(_plain(dataset.cell(r, column)))) for r in dataset.rows[:200]]
        )
        sheet.column_dimensions[get_column_letter(index)].width = min(max(widest + 2, 10), 48)

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _excel_datetime(value: Any) -> Any:
    """Convert a stored UTC datetime to a naive local one for Excel.

    Two reasons this is the right boundary rather than a workaround: Excel
    rejects timezone-aware datetimes outright, and Part II §1 says UTC is what
    is *stored* while local time is a presentation concern — an export is
    presentation.
    """
    if isinstance(value, dt.datetime):
        from app.core.i18n import to_local

        return to_local(value).replace(tzinfo=None)
    if isinstance(value, dt.date):
        return value
    return _plain(value)


def _write_cell(cell, value: Any, column: Column) -> None:
    """Write a typed cell so Excel can sum and sort it."""
    if value is None or value == "":
        cell.value = ""
        return

    if column.kind == "money":
        cell.value = float(value) if not isinstance(value, str) else value
        # JOD is a three-decimal currency (fils).
        cell.number_format = "#,##0.000"
    elif column.kind == "number":
        cell.value = value if isinstance(value, (int, float)) else _plain(value)
        cell.number_format = "#,##0"
    elif column.kind == "date":
        cell.value = _excel_datetime(value)
        cell.number_format = (
            "yyyy-mm-dd hh:mm" if isinstance(cell.value, dt.datetime) else "yyyy-mm-dd"
        )
    else:
        cell.value = _plain(value)


# ---------------------------------------------------------------------------
# HTML / PDF
# ---------------------------------------------------------------------------


def to_html(request: Request, dataset: Dataset) -> str:
    """Render the print-ready HTML used for both PDF and browser printing."""
    from app.core.templating import templates

    response = templates.TemplateResponse(
        request, "admin/exports/report.html", {"dataset": dataset}
    )
    return response.body.decode("utf-8")


def to_pdf(request: Request, dataset: Dataset) -> bytes:
    if not pdf_available():
        raise PdfUnavailable()

    from weasyprint import HTML

    return HTML(string=to_html(request, dataset), base_url=str(request.base_url)).write_pdf()


class PdfUnavailable(AppError):
    """PDF rendering is not possible in this environment.

    Deliberately explicit rather than a silent fallback: an accountant who asked
    for a PDF should be told to use the print view, not handed a CSV.
    """

    status_code = 503
    code = "pdf_unavailable"
    message = (
        "PDF export needs WeasyPrint's system libraries, which are not installed "
        "here. Use the print view and choose 'Save as PDF'."
    )


# ---------------------------------------------------------------------------
# Response
# ---------------------------------------------------------------------------


def export_response(request: Request, dataset: Dataset, fmt: str) -> Response:
    """Build the download response for the requested format."""
    if fmt == "csv":
        payload, media = to_csv(dataset), "text/csv; charset=utf-8"
    elif fmt in {"xlsx", "excel"}:
        fmt = "xlsx"
        payload = to_xlsx(dataset)
        media = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    elif fmt == "pdf":
        payload, media = to_pdf(request, dataset), "application/pdf"
    elif fmt == "html":
        # The print view: rendered inline, not downloaded.
        return Response(to_html(request, dataset), media_type="text/html")
    else:
        raise AppError(
            "Unsupported export format.", code="unsupported_format", status_code=400
        )

    log.info(
        "report_exported",
        extra={"report": dataset.title, "format": fmt, "rows": len(dataset.rows)},
    )
    return StreamingResponse(
        io.BytesIO(payload),
        media_type=media,
        headers={
            "Content-Disposition": content_disposition(dataset, fmt),
            "Content-Length": str(len(payload)),
        },
    )
