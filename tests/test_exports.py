"""Report exports (Part I §2.2).

§2.2: *"Every report and dashboard view must be exportable (CSV, Excel, and PDF
at minimum) for offline use and accounting handoff."*

Arabic is what makes this non-trivial, so most of what is pinned here is about
Arabic surviving the round trip — the BOM Excel needs, the RTL sheet direction,
and the filename encoding an HTTP header would otherwise reject.
"""

from __future__ import annotations

import datetime as dt
import io
from decimal import Decimal

import pytest

from app.services.exports import (
    Column,
    Dataset,
    content_disposition,
    to_csv,
    to_xlsx,
)
from tests.test_checkout import db  # noqa: F401 - fixture


@pytest.fixture
def dataset() -> Dataset:
    return Dataset(
        title="Sales",
        language="en",
        columns=[
            Column("number", "Order"),
            Column("placed", "Placed", kind="date"),
            Column("qty", "Qty", kind="number"),
            Column("total", "Total", kind="money"),
        ],
        rows=[
            {
                "number": "JEC-260812-001",
                "placed": dt.datetime(2026, 8, 12, 9, 30, tzinfo=dt.timezone.utc),
                "qty": 2,
                "total": Decimal("26.640"),
            },
            {
                "number": "JEC-260812-002",
                "placed": dt.datetime(2026, 8, 12, 14, 5, tzinfo=dt.timezone.utc),
                "qty": 1,
                "total": Decimal("14.800"),
            },
        ],
        meta={"Period": "August 2026"},
        totals={"number": "Total", "qty": 3, "total": Decimal("41.440")},
    )


@pytest.fixture
def arabic_dataset() -> Dataset:
    return Dataset(
        title="المخزون",
        language="ar",
        columns=[Column("product", "المنتج"), Column("qty", "الكمية", kind="number")],
        rows=[{"product": "الكتاب المقدس", "qty": 5}],
    )


# ---------------------------------------------------------------------------
# CSV
# ---------------------------------------------------------------------------


def test_csv_starts_with_a_utf8_bom(arabic_dataset: Dataset):
    """Without the BOM, Excel on Windows opens Arabic as mojibake."""
    assert to_csv(arabic_dataset).startswith(b"\xef\xbb\xbf")


def test_csv_round_trips_arabic(arabic_dataset: Dataset):
    text = to_csv(arabic_dataset).decode("utf-8-sig")
    assert "المخزون" in text
    assert "الكتاب المقدس" in text


def test_csv_contains_headers_rows_and_totals(dataset: Dataset):
    text = to_csv(dataset).decode("utf-8-sig")
    assert "Order" in text and "Total" in text
    assert "JEC-260812-001" in text
    assert "41.440" in text, "the totals row is written"
    assert "August 2026" in text, "filter context travels with the export"


def test_csv_neutralises_formula_injection():
    """An exported note starting with '=' must be text, not a calculation."""
    payload = to_csv(
        Dataset(
            title="Notes",
            columns=[Column("note", "Note")],
            rows=[{"note": "=1+1"}, {"note": "+cmd"}, {"note": "safe"}],
        )
    ).decode("utf-8-sig")

    assert "'=1+1" in payload
    assert "'+cmd" in payload
    assert "'safe" not in payload, "ordinary values are left alone"


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------


def _sheet(dataset: Dataset):
    from openpyxl import load_workbook

    return load_workbook(io.BytesIO(to_xlsx(dataset))).active


def test_xlsx_is_a_valid_workbook(dataset: Dataset):
    assert to_xlsx(dataset)[:2] == b"PK"


def test_xlsx_writes_numbers_as_numbers(dataset: Dataset):
    """A total must sum in the spreadsheet, not sit there as a string."""
    sheet = _sheet(dataset)
    values = [c.value for row in sheet.iter_rows() for c in row]
    assert 26.64 in values or Decimal("26.640") in values
    assert 2 in values


def test_xlsx_strips_timezones_from_datetimes(dataset: Dataset):
    """Excel rejects aware datetimes; the export converts to naive local time."""
    sheet = _sheet(dataset)
    stamps = [
        c.value
        for row in sheet.iter_rows()
        for c in row
        if isinstance(c.value, dt.datetime)
    ]
    assert stamps, "the date column was written"
    assert all(s.tzinfo is None for s in stamps)
    # 09:30 UTC is 12:30 in Amman (UTC+3, no DST).
    assert any(s.hour == 12 and s.minute == 30 for s in stamps)


def test_xlsx_is_right_to_left_for_arabic(arabic_dataset: Dataset):
    """The columns must read in the right order, not merely contain Arabic."""
    assert _sheet(arabic_dataset).sheet_view.rightToLeft is True
    assert _sheet(Dataset(title="X", columns=[Column("a", "A")], rows=[])).sheet_view.rightToLeft is False


def test_xlsx_sheet_name_is_sanitised():
    """Excel forbids some characters and caps the name at 31 chars."""
    sheet = _sheet(
        Dataset(title="A/B:C" + "x" * 40, columns=[Column("a", "A")], rows=[])
    )
    assert len(sheet.title) <= 31
    assert not set(sheet.title) & set("\\/*?:[]")


def test_xlsx_freezes_the_header(dataset: Dataset):
    assert _sheet(dataset).freeze_panes is not None


# ---------------------------------------------------------------------------
# Download headers
# ---------------------------------------------------------------------------


def test_arabic_filename_uses_rfc5987(arabic_dataset: Dataset):
    """HTTP headers are latin-1; an Arabic name needs filename*."""
    header = content_disposition(arabic_dataset, "csv")

    assert "filename*=UTF-8''" in header
    assert "%D8" in header, "the Arabic name is percent-encoded"
    # And it must still encode as latin-1, or Starlette raises on send.
    header.encode("latin-1")


def test_ascii_fallback_filename_is_present(arabic_dataset: Dataset):
    header = content_disposition(arabic_dataset, "csv")
    assert 'filename="' in header
    assert header.split('filename="')[1].split('"')[0].endswith(".csv")


# ---------------------------------------------------------------------------
# PDF degradation
# ---------------------------------------------------------------------------


def test_pdf_reports_unavailability_explicitly():
    """An accountant who asked for PDF is told why, not handed a CSV."""
    from app.services.exports import PdfUnavailable, pdf_available

    if pdf_available():
        pytest.skip("WeasyPrint natives are present in this environment")

    error = PdfUnavailable()
    assert error.status_code == 503
    assert error.code == "pdf_unavailable"
    assert "print view" in error.message


# ---------------------------------------------------------------------------
# The reports §2.2 names by name
# ---------------------------------------------------------------------------


def test_every_report_key_builds_in_both_languages(db):
    """§2.2 lists the reports the shop must be able to export: "sales,
    inventory, consignment, money boxes, promocodes, returns, staff activity".

    Promocodes, returns and staff activity had no dataset at all, so three of
    the seven named reports simply did not exist. This walks the whole registry
    rather than the three, so the next one added cannot be half-wired either.
    """
    import datetime as dt

    from app.services import report_datasets

    today = dt.date.today()
    for key in report_datasets.REPORT_KEYS:
        for language in ("en", "ar"):
            dataset = report_datasets.build(
                db,
                key,
                language=language,
                start_date=today - dt.timedelta(days=30),
                end_date=today,
            )
            assert dataset.columns, f"{key} [{language}] has no columns"
            assert dataset.title, f"{key} [{language}] has no title"
            # A label falling back to its own key is how an untranslated column
            # header reaches a customer-facing PDF.
            untranslated = [c.label for c in dataset.columns if "." in c.label]
            assert not untranslated, f"{key} [{language}] raw keys: {untranslated}"


def test_the_period_window_includes_its_own_last_day(db):
    """A half-open window ending at midnight *on* the end date drops everything
    that happened that day — a report's most recent and most scrutinised day."""
    import datetime as dt

    from app.services.report_datasets import _period_bounds

    start, end = _period_bounds(dt.date(2026, 1, 1), dt.date(2026, 1, 31))
    assert start == dt.datetime(2026, 1, 1, tzinfo=dt.timezone.utc)
    assert end == dt.datetime(2026, 2, 1, tzinfo=dt.timezone.utc)
